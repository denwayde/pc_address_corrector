import requests
import json
from dadata import Dadata
from dotenv import load_dotenv
import os
import string
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import httpx
import time
from openpyxl.utils.exceptions import InvalidFileException
import asyncio
# Загружаем переменные из файла .env
load_dotenv()
# Пример использования переменных
token = os.getenv("TOKEN")
secret = os.getenv("SECRET")
ya_apikey = os.getenv("YA_APIKEY")

border_style = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font_style = Font(bold=False, size=10)
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def is_punctuation(char):
    return char in string.punctuation

def cut_all_tire(text_from_exel):
    if is_punctuation(text_from_exel[0]):
        text_from_exel = text_from_exel[1: ]
        text_from_exel = text_from_exel.replace("-", " ")
    return text_from_exel


async def dadata_proccess(req): 
    try:#"Уфа г, Сун-Ят-Сена ул, д.11, кв.138"
        dadata = Dadata(token, secret)
        final_result = ''
        result = dadata.suggest("address", req)
        if result != []:
            #print(result[0]['unrestricted_value'])
            final_result = result[0]['unrestricted_value']
        await asyncio.sleep(0)
        return final_result   
    except httpx.ConnectTimeout:
        print('Прерыв соединения с удаленным сервером')
        #return "Ошибка при обработке запроса" 
        
def correct_location(s1):
    m = re.findall(r'[а-яa-z][А-Я]', s1)
    if m != []:
        for x in m:
            r = x[0]+". "+x[1]
            #print(r)
            s1 = s1.replace(x, r)
    return s1


def correct_street(s1):
    m = re.findall(r'\d\/[а-яА-Яa-zA-Z]', s1)
    if m != []:
        for x in m:
            r = x.replace("/", "")
            s1 = s1.replace(x, r)
    return s1


async def ya_maps_proccess(req):
    r = requests.get(f'https://suggest-maps.yandex.ru/v1/suggest?apikey={ya_apikey}&text={req}&print_address=1')
    final_result = ''
    if hasattr(json.loads(r.text), 'results') or 'results' in json.loads(r.text):
        final_result = json.loads(r.text)["results"][0]['address']['formatted_address']
    return final_result 



async def result_from_apis(req):###-----------------------chto to ne to imenno zdes
    res = await dadata_proccess(req)
    if res == []:
        res1 = await dadata_proccess(cut_all_tire(req))
        if res1 == []:
            res3 = await ya_maps_proccess(req)
            if res3 == []:
                #print("res3 from ya api" + dadata_proccess(ya_maps_proccess(cut_all_tire(req))))
                return dadata_proccess(ya_maps_proccess(cut_all_tire(req)))
            else:
                #print("samii finalnii res"+ dadata_proccess(res3))
                return dadata_proccess(res3)
        else:
            #print("eto vtoroi res"+ dadata_proccess(req))
            return res1
    else:
        #print("eto pervii result"+ dadata_proccess(cut_all_tire(req)))
        return res



def this_phone_num(text_from_exel):
    final_result = True
    digit = 0
    for v in text_from_exel:
        if v.isdigit():
            digit = digit + 1
    if digit == 0 or digit >= 11:
        final_result = False
    return final_result



async def excel_procces(file_name, dispaly_text):
    try:
        wb = load_workbook(file_name)#'Arc.xlsx'
        # Выбираем активный лист
        sheet = wb.active
        sheet.column_dimensions['H'].width = 80
        for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2): 
            try:
                value_a = row[0]
                value_a = correct_street(correct_location(value_a))#OBRABOTKA STROKI V YACHEIKE A: IZBAVLYAEMSYA OT KRIVIH OBOZNACHENII SELA I GORODA, I KRIVIH OBOZNACHENII DOMA
                cell = sheet.cell(row=idx, column=8)#POLUCHAEM YACHEIKI IZ STOLBA h
                # print(cell.value)
                if cell.value == None or cell.value == '' or cell.value == 'По непонятным причинам, не удалось нормализовать адрес.':
                    if this_phone_num(value_a):
                        exression = await result_from_apis(value_a)
                        if exression != '':
                            await dispaly_text(exression)
                            sheet.cell(row = idx, column=8, value = exression)
                            fill = PatternFill(start_color="BFEA7C", end_color="BFEA7C", fill_type="solid")
                            cell.fill = fill
                        else:
                            sheet.cell(row = idx, column=8, value = 'По непонятным причинам, не удалось нормализовать адрес.')
                            fill = PatternFill(start_color="F6F193", end_color="F6F193", fill_type="solid")
                            cell.fill = fill   
                    else:
                        sheet.cell(row = idx, column=8, value = f"В ячейке A[{idx}] не адрес")
                        # Создаем объект, представляющий стиль с заполнением ячейки
                        fill = PatternFill(start_color="FFB996", end_color="FFB996", fill_type="solid")
                        cell.fill = fill
                else:
                    await dispaly_text(f"В ячейке H[{idx}] запись есть")
                    #pass
                cell.border = border_style
                cell.font = font_style
            except InvalidFileException as e:
                await dispaly_text(f'Ошибка при обращении к файлу: {e}')
            except IndexError:
                await dispaly_text("Vishli za predeli")
            except Exception as e:
                await dispaly_text(f'Произошла непредвиденная ошибка: {e}')
            await asyncio.sleep(1)
        sheet.cell(row=1, column=8, value='Адреса').font = Font(bold=True)
        sheet.cell(row=1, column=8, value='Адреса').border = border_style
        wb.save(file_name)       
        await dispaly_text("Запись завершена")
    except asyncio.CancelledError:
        await dispaly_text("Отмена...")



#excel_procces('Arc.xlsx', print)


