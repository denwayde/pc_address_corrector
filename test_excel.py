from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

border_style = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
#border_style1 = Border(right=Side())
# Задаем стиль для шрифта (не жирный)
font_style = Font(bold=False, size=10)
# Задаем заливку для ячейки
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# cell.fill = fill
# Открываем Excel файл
wb = load_workbook('arhive.xlsx')

# Выбираем активный лист
sheet = wb.active
sheet.column_dimensions['H'].width = 60
# Проходим по всем строкам в столбце A и выводим ячейки с записями
# for cell in sheet['A']:
#     if cell.value:
#         print(cell.row)
        #print(cell.value)
# Закрываем файл Excel
#wb.close()
# Проход по значениям столбца A, обработка и запись в столбец H
for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):  # Начинаем с 2 строки, чтобы пропустить заголовок
    value_a = row[0]
    # Обработка значения из столбца A
    processed_value = "Provereno " + str(value_a)  # Например, переводим значение в верхний регистр
    # Запись обработанного значения в столбец H
    sheet.cell(row=idx, column=8, value=processed_value)

    # Применяем стиль обводки к ячейке, которую ты установил ранее
    cell = sheet.cell(row=idx, column=8)
    cell.border = border_style
    cell.font = font_style
sheet.cell(row=1, column=8, value='Адреса').font = Font(bold=True)
sheet.cell(row=1, column=8, value='Адреса').border = border_style

# Сохранение изменений в файл
wb.save('arhive.xlsx')
