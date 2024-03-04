# -*- coding: utf-8 -*-
import requests
import json
from dadata import Dadata
from dotenv import load_dotenv
import os
from db_func import delete_or_insert_data, select_data
import time
import string
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side


# Загружаем переменные из файла .env
load_dotenv()
# Пример использования переменных
token = os.getenv("TOKEN")
secret = os.getenv("SECRET")
ya_apikey = os.getenv("YA_APIKEY")



def substr_right_proccess(substr_right):
    
    if substr_right != '':
        if not substr_right[0].isalpha():
            substr_right = substr_right[1: ]
            
            if "кв" in substr_right.lower():
                substr_right = ", "+ substr_right
            else:
                substr_right = ", кв "+ substr_right
        else:
            substr_right = ", "+ substr_right
        substr_right = substr_right.replace("  ", " ")
    else:
        substr_right = ''
    return substr_right 


def is_punctuation(char):
    return char in string.punctuation

def check_none(text_from_respond):
    if text_from_respond == None:
        text_from_respond = ''
    return text_from_respond

def is_none(text_from_respond):
    if text_from_respond == None:
        return False
    return True

import re


def address_proccess(text_from_exel):
    dadata = Dadata(token, secret)
    digit = 0
    result = None
    
    #dadata_client = dadata.Dadata(timeout=10)
    result = dadata.suggest("address", text_from_exel)
    
    if result != []:
        final_result = ''
        #print(result[0])
        value = result[0]['value']
        postal_code = check_none(result[0]['data']['postal_code'])
        region = check_none(result[0]['data']['region_with_type'])
        area = check_none(result[0]['data']['area_with_type'])
        city = check_none(result[0]['data']['city_with_type'])
        settle = check_none(result[0]['data']['settlement_with_type'])
        street = check_none(result[0]['data']['street_with_type'])
        house = check_none(result[0]['data']['house'])

        index = text_from_exel.find(house)
        substr_right = text_from_exel[index+len(house): ]
        house_result_index = value.find(house)
        substr_result_left = value[ :house_result_index+len(house)]
        flat = substr_right_proccess(substr_right)
        final_result = postal_code + ", "+ region +", "+ area +", "+ city +", "+ settle +", "+ street +", "+ house + flat + right_end
        final_result = final_result.replace(" ,", "")
        dadata.close()
        return final_result
    else:
        #print("Netu rezultata poiska v dadata")
        r = requests.get(f'https://suggest-maps.yandex.ru/v1/suggest?apikey={ya_apikey}&text={text_from_exel}&print_address=1')
        if hasattr(json.loads(r.text), 'results') or 'results' in json.loads(r.text):
            #print(json.loads(r.text)["results"][0])
            check_location_arr = {}
            house = ''
            for x in json.loads(r.text)["results"][0]['address']['component']:
                if x['kind'] == ['LOCALITY']:
                    check_location_arr['local'] = x['name']
                if x['kind'] == ['STREET']:
                    st = ''
                    for z in x['name'].split(' '):
                        if z.istitle():
                            st = st+" "+z 
                    st = st.lstrip()
                    check_location_arr['street'] = st
                if x['kind'] == ['HOUSE']:
                    house = x['name']
            #print(check_location_arr)
            house_count = text_from_exel.count(house)
            if house_count == 0:
                street_index = text_from_exel.find(check_location_arr['street'])
                substr_right = text_from_exel[street_index+len(check_location_arr['street']) : ] 
                #print(s13[street_index+len(check_location_arr['street']) : ])
                substr_right_right = ''
                for c in substr_right:
                    if is_punctuation(c):
                        punct_index = substr_right.find(c)
                        substr_right_right = substr_right[punct_index + len(c) : ]
                        substr_right_right = ", " + substr_right_right
                        substr_right_right = substr_right_right.replace("  ", " ")
                        break
                dadata = Dadata(token=token, secret=secret)
                result1 = dadata.suggest("address", check_location_arr['local']+" "+check_location_arr['street']+ " " + house)
                if result1 != []:
                    final_result = result1[0]['data']['postal_code']+", "+result1[0]['data']['region']+", "+result1[0]['data']['city_with_type']+", "+result1[0]['data']['street_with_type']+", "+substr_right_right
                    time.sleep(1)
                    dadata.close()
                    return final_result
                else:
                    final_result = "Невозможно преобразовать адрес. Error: in home length = 0"
                    time.sleep(1)
                    return final_result
            else:
                if hasattr(check_location_arr, 'local') and hasattr(check_location_arr, 'street'):
                    if check_location_arr['local'] in text_from_exel and check_location_arr['street'] in text_from_exel:
                        #house = json.loads(r.text)["results"][0]['address']['component'][-2]['name']
                        index = text_from_exel.find(house)
                        substr_right = text_from_exel[index + len(house): ]
                        flat = substr_right_proccess(substr_right)
                        formatted_address = json.loads(r.text)["results"][0]['address']['formatted_address'] + flat
                        time.sleep(1)
                        address_proccess(formatted_address)
                    else:
                        final_result = "Невозможно преобразовать адрес. Error is_valid_location False"
                        dadata.close()
                        return final_result
                else:
                    final_result = "Невозможно преобразовать адрес. Error is_valid_location False"
                    dadata.close()
                    return final_result
        else:
            final_result = "Невозможно преобразовать адрес. Error No results in yandex maps neither in dadata"
            dadata.close()
            time.sleep(1)
            return final_result

border_style = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font_style = Font(bold=False, size=10)
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
wb = load_workbook('Arc.xlsx')
# Выбираем активный лист
sheet = wb.active
sheet.column_dimensions['H'].width = 60

address_array = []
for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):  # Начинаем с 2 строки, чтобы пропустить заголовок
    value_a = row[0]
    address_array.append(value_a)
    # Обработка значения из столбца A
    #processed_value = "Provereno " + str(value_a)  
    # Запись обработанного значения в столбец H
    #sheet.cell(row=idx, column=8, value=address_proccess(str(value_a)))# ------ ZADAEM ZNACHENIE V YACHEIKU -------------------------
    # Применяем стиль обводки к ячейке, которую ты установил ранее
    # cell = sheet.cell(row=idx, column=8)
    # cell.border = border_style
    # cell.font = font_style


# Сохранение изменений в файл
wb.save('Arc.xlsx')

correct_address_arr = []
for n in address_array:
    correct_address_arr.append(address_proccess(n))


wb = load_workbook('Arc.xlsx')
# Выбираем активный лист
sheet = wb.active

for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
    try:
        sheet.cell(row=idx, column=8, value=correct_address_arr[idx])
    except IndexError:
        print("Vishli za predeli")
    cell = sheet.cell(row=idx, column=8)
    cell.border = border_style
    cell.font = font_style

sheet.cell(row=1, column=8, value='Адреса').font = Font(bold=True)
sheet.cell(row=1, column=8, value='Адреса').border = border_style
wb.save('Arc.xlsx')
# cell.fill = fill
#sss1 = "ост пункт 1680 км, Центральная ул., 1, 1"
#sss = ", 450071, Уфа г, 50 лет СССР ул, д.40, кв.33"#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#print(address_proccess(sss2))

