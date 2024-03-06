import requests
import json
from dadata import Dadata
from dotenv import load_dotenv
import os
import string
import re
from openpyxl import load_workbook
import httpx
from openpyxl.styles import Font, PatternFill, Border, Side
# Загружаем переменные из файла .env
load_dotenv()
# Пример использования переменных
token = os.getenv("TOKEN")
secret = os.getenv("SECRET")
ya_apikey = os.getenv("YA_APIKEY")



def is_punctuation(char):
    return char in string.punctuation

def check_none(text_from_respond):
    if text_from_respond == None or len(text_from_respond)==0:
        text_from_respond = ''
    return text_from_respond

def is_none(text_from_respond):
    if text_from_respond == None:
        return False
    return True

border_style = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font_style = Font(bold=False, size=10)
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")



def this_phone_num(text_from_exel):
    final_result = True
    for v in text_from_exel:
        if v.isdigit():
            digit = digit + 1
    if digit == 0 or digit >= 11:
        final_result = False
    return final_result

def cut_all_tire(text_from_exel):
    if is_punctuation(text_from_exel[0]):
        text_from_exel = text_from_exel[text_from_exel.find(text_from_exel[0])+len(text_from_exel[0]): ]
        text_from_exel = text_from_exel.replace("-", " ")
    return text_from_exel


def addr_right_proccess(addrr):
    procceed=''
    mtc = re.search(r'\d+[\,\s\-\.\b]?', addrr)
    #print("mtc---------------" + str(mtc))
    if mtc != None:
        #print("mtc.group()---------------" + str(mtc.group()))
        addr_right = addrr[addrr.find(mtc.group())+len(mtc.group()) : ]#poluchili to chto sprava ot naidennogo chisla
        #print("addr_right---------------"+str(addr_right))
        return addr_right_proccess(addr_right)
    else:
        if addrr != '':
            procceed = ', ' + addrr
            procceed = procceed.replace("  ", " ")
        else:
            procceed = addrr
        return procceed

def right_end_proccess(text_from_exel):
    right_end = addr_right_proccess(text_from_exel)
    if right_end != '':
        text_from_exel = text_from_exel[ :text_from_exel.find(right_end)]
    return right_end


def address_proccess(text_from_exel):
    dadata = Dadata(token, secret)
    final_result = ''
    try:
        result = dadata.suggest("address", text_from_exel)
        if result !=[]:
            # print(result[0]['value'])
            # excel_procces()
            final_result = result[0]['value']
        return final_result
    except httpx.ConnectTimeout:
        print("Не удалось")
        

def excel_procces(file_name, correct_address):
    wb = load_workbook(file_name)#'Arc.xlsx'
    # Выбираем активный лист
    sheet = wb.active
    sheet.column_dimensions['H'].width = 60
    for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2): 
        try:
            value_a = row[0]
            if this_phone_num(value_a):
                correct_value = address_proccess()
            else:
                correct_value = "Это точно не адрес"
        except IndexError:
            print("Vishli za predeli")
            correct_value = "вышли за пределы"
        finally:
            sheet.cell(row = idx, column=8, value = correct_address(correct_value))
            cell = sheet.cell(row=idx, column=8)
            cell.border = border_style
            cell.font = font_style

    sheet.cell(row=1, column=8, value='Адреса').font = Font(bold=True)
    sheet.cell(row=1, column=8, value='Адреса').border = border_style
    wb.save(file_name)


#excel_procces('Arc.xlsx', address_proccess)
dadata = Dadata(token, secret)
    
result = dadata.suggest("address", "Уфа г, Сельская Богородская ул, д.45, кв. 69")

print(result[0])